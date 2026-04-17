"""Generate Sprint Planning Excel - All 15 sprints with detailed task sheets."""
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

wb = Workbook()

# ── Styles ────────────────────────────────────────────────────────────────
HEADER_BG = PatternFill("solid", fgColor="1A1A2E")
HEADER_FG = Font(name="Arial", bold=True, color="FFFFFF", size=11)
NORMAL = Font(name="Arial", size=10)
BOLD = Font(name="Arial", bold=True, size=10)
SMALL = Font(name="Arial", size=9, color="666666")
STATUS_DONE = PatternFill("solid", fgColor="C8E6C9")
STATUS_IP = PatternFill("solid", fgColor="FFF9C4")
STATUS_TODO = PatternFill("solid", fgColor="FFCDD2")
TOTAL_BG = PatternFill("solid", fgColor="F5F5F5")
thin = Border(
    left=Side(style="thin", color="CCCCCC"), right=Side(style="thin", color="CCCCCC"),
    top=Side(style="thin", color="CCCCCC"), bottom=Side(style="thin", color="CCCCCC"),
)
ctr = Alignment(horizontal="center", vertical="center", wrap_text=True)
lft = Alignment(horizontal="left", vertical="center", wrap_text=True)

SPRINT_COLORS = {
    1: "1565C0", 2: "2E7D32", 3: "E65100", 4: "6A1B9A",
    5: "AD1457", 6: "00838F", 7: "F9A825", 8: "558B2F",
    9: "BF360C", 10: "283593", 11: "4E342E", 12: "37474F",
    13: "FF5722", 14: "9C27B0", 15: "2196F3",
}
TAB_COLORS = {
    1: "42A5F5", 2: "66BB6A", 3: "FFA726", 4: "AB47BC",
    5: "EC407A", 6: "26C6DA", 7: "FFEE58", 8: "9CCC65",
    9: "FF7043", 10: "5C6BC0", 11: "8D6E63", 12: "78909C",
    13: "FF5722", 14: "9C27B0", 15: "2196F3",
}

def set_col_widths(ws, widths):
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

def title_row(ws, text, color, cols=7):
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=cols)
    ws["A1"] = text
    ws["A1"].font = Font(name="Arial", bold=True, size=14, color="FFFFFF")
    ws["A1"].fill = PatternFill("solid", fgColor=color)
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 38

def header_row(ws, row, headers):
    for i, h in enumerate(headers, 1):
        c = ws.cell(row=row, column=i, value=h)
        c.font = HEADER_FG
        c.fill = HEADER_BG
        c.alignment = ctr
        c.border = thin
    ws.row_dimensions[row].height = 28

def data_row(ws, row, data, wrap_cols=None):
    wrap_cols = wrap_cols or []
    for i, val in enumerate(data, 1):
        c = ws.cell(row=row, column=i, value=val)
        c.font = NORMAL
        c.border = thin
        c.alignment = lft if i in wrap_cols else ctr
    ws.row_dimensions[row].height = 28

def status_fill(status):
    s = status.upper()
    if s == "DONE":
        return STATUS_DONE
    if s in ("IN PROGRESS", "IP"):
        return STATUS_IP
    return STATUS_TODO

def build_sprint_sheet(ws, sprint_num, sprint_title, weeks, color, tasks):
    ws.sheet_properties.tabColor = TAB_COLORS[sprint_num]
    set_col_widths(ws, [5, 50, 14, 10, 10, 40, 8])
    title_row(ws, f"SPRINT {sprint_num}: {sprint_title} ({weeks})", color)
    ws.merge_cells("A2:G2")
    ws["A2"] = f"{len(tasks)} tasks"
    ws["A2"].font = SMALL
    ws["A2"].alignment = Alignment(horizontal="center")
    header_row(ws, 4, ["#", "Task", "Category", "Pri", "Status", "Notes / Details", "Est."])
    for idx, t in enumerate(tasks):
        r = 5 + idx
        data_row(ws, r, t, wrap_cols=[2, 6])
        ws.cell(row=r, column=5).fill = status_fill(t[4])
        ws.cell(row=r, column=5).font = BOLD


# ══════════════════════════════════════════════════════════════════════════
# SHEET 1: Sprint Overview (master)
# ══════════════════════════════════════════════════════════════════════════
ws1 = wb.active
ws1.title = "Overview"
ws1.sheet_properties.tabColor = "1A1A2E"
set_col_widths(ws1, [5, 30, 12, 12, 12, 12, 8, 44, 10])

ws1.merge_cells("A1:I1")
ws1["A1"] = "KUWAIT WHATSAPP GROWTH ENGINE - FULL SPRINT PLAN (S1-S15)"
ws1["A1"].font = Font(name="Arial", bold=True, size=16, color="FFFFFF")
ws1["A1"].fill = PatternFill("solid", fgColor="1A1A2E")
ws1["A1"].alignment = Alignment(horizontal="center", vertical="center")
ws1.row_dimensions[1].height = 42

ws1.merge_cells("A2:I2")
ws1["A2"] = "15 Sprints | 30 Weeks | 29 DB Tables | ~105 API Endpoints | 132 Source Files | 17 Services"
ws1["A2"].font = Font(name="Arial", size=11, color="666666", italic=True)
ws1["A2"].alignment = Alignment(horizontal="center")
ws1.row_dimensions[2].height = 22

header_row(ws1, 4, ["#", "Sprint Name", "Weeks", "Duration", "Phase", "Status", "Tasks", "Key Deliverables", "Priority"])

sprints_overview = [
    [1, "Foundation & Auth", "W1-W2", "2 wks", "Phase 1", "DONE", 12, "JWT, RBAC, Docker 7 svc, PostgreSQL RLS, Redis, Celery, Next.js 15", "P0"],
    [2, "Core CRM", "W3-W4", "2 wks", "Phase 2", "DONE", 9, "Contacts CRUD, tags, custom fields, CSV import, search, data table", "P0"],
    [3, "WhatsApp Integration", "W5-W6", "2 wks", "Phase 3", "DONE", 10, "Cloud API provider, webhooks, conversations, messages, WebSocket inbox", "P0"],
    [4, "Smart Routing & Automation", "W7-W8", "2 wks", "Phase 4", "DONE", 9, "4-factor routing, automation rules (7 ops, 8 actions), Celery execution", "P0"],
    [5, "AI Dialect Engine", "W9-W10", "2 wks", "Phase 5", "DONE", 8, "Kuwaiti dialect detection, intent (10 types), sentiment, Claude API", "P1"],
    [6, "Sales Pipeline", "W11-W12", "2 wks", "Phase 6", "DONE", 8, "Pipeline CRUD, kanban board, deals, stages, activity log, drag-drop", "P0"],
    [7, "Payments & Billing", "W13-W14", "2 wks", "Phase 7", "DONE", 9, "Tap K-Net/Visa/MC, subscriptions, invoices, plan limits, webhook", "P0"],
    [8, "Shipping Integration", "W15-W16", "2 wks", "Phase 8", "DONE", 8, "Carrier abstraction, Aramex, tracking events, WhatsApp notifications", "P1"],
    [9, "Landing Pages", "W17-W18", "2 wks", "Phase 9", "DONE", 8, "Block editor (8 types), publish, public pages, conversion tracking", "P1"],
    [10, "Analytics Dashboard", "W19-W20", "2 wks", "Phase 10", "DONE", 6, "Dashboard KPIs, message/pipeline/team/LP/automation stats", "P1"],
    [11, "Compliance & Audit", "W21-W22", "2 wks", "Phase 11", "DONE", 7, "Audit logs, 8-check compliance, data residency report, security list", "P0"],
    [12, "Optimization & Hardening", "W23-W24", "2 wks", "Phase 12", "DONE", 7, "Redis cache, Prometheus metrics, 5 periodic tasks, Alembic, .env", "P1"],
    [13, "Production Readiness", "W25-W26", "2 wks", "Deploy", "TODO", 18, "Secrets, SSL, AWS me-south-1, CI/CD, RDS, load test, DNS, monitoring", "P0"],
    [14, "UI Polish & i18n", "W27-W28", "2 wks", "Frontend", "TODO", 14, "shadcn/ui, Arabic RTL, translations, responsive, dark mode, toasts", "P1"],
    [15, "Testing & QA", "W29-W30", "2 wks", "Quality", "TODO", 12, "Unit tests (7 services), integration (3 flows), OWASP security audit", "P0"],
]

for idx, row in enumerate(sprints_overview):
    r = 5 + idx
    ws1.row_dimensions[r].height = 34
    data_row(ws1, r, row, wrap_cols=[8])
    ws1.cell(row=r, column=6).fill = status_fill(row[5])
    ws1.cell(row=r, column=6).font = BOLD
    # Color code sprint number
    ws1.cell(row=r, column=1).fill = PatternFill("solid", fgColor=TAB_COLORS[row[0]])
    ws1.cell(row=r, column=1).font = Font(name="Arial", bold=True, size=10, color="FFFFFF")

# Totals
r_sum = 5 + len(sprints_overview)
ws1.merge_cells(f"A{r_sum}:D{r_sum}")
ws1.cell(row=r_sum, column=1, value="TOTAL (15 SPRINTS, 30 WEEKS)").font = BOLD
ws1.cell(row=r_sum, column=7, value=f"=SUM(G5:G{r_sum-1})")
ws1.cell(row=r_sum, column=7).font = BOLD
for c in range(1, 10):
    ws1.cell(row=r_sum, column=c).border = thin
    ws1.cell(row=r_sum, column=c).fill = TOTAL_BG


# ══════════════════════════════════════════════════════════════════════════
# SHEETS 2-16: Individual Sprint Details (S1-S15)
# ══════════════════════════════════════════════════════════════════════════

all_sprint_tasks = {
    1: ("FOUNDATION & AUTH", "W1-W2", [
        [1, "Set up project structure (backend/frontend/docker)", "DevOps", "P0", "DONE", "Monorepo with backend/, frontend/, docker/", "2h"],
        [2, "Create FastAPI app factory with lifespan", "Backend", "P0", "DONE", "main.py: middleware stack, error handlers", "3h"],
        [3, "Set up SQLAlchemy 2.0 async + Alembic", "Backend", "P0", "DONE", "Base, UUIDMixin, TimestampMixin, TenantMixin", "3h"],
        [4, "Create User, Role, UserRole, APIKey models", "Backend", "P0", "DONE", "auth.py: smart routing fields on User", "3h"],
        [5, "Create Company model with encrypted config", "Backend", "P0", "DONE", "WhatsApp, Twilio, Tap config fields", "2h"],
        [6, "Build JWT auth (access 15m + refresh 7d)", "Backend", "P0", "DONE", "security.py: bcrypt 12 rounds, HS256", "4h"],
        [7, "Build auth service (register, login, refresh, me)", "Backend", "P0", "DONE", "auth_service.py: 4 endpoints", "4h"],
        [8, "Create RBAC (5 roles, 30+ permissions)", "Backend", "P0", "DONE", "permissions.py: ROLE_PERMISSIONS mapping", "2h"],
        [9, "Create Docker Compose (7 services)", "DevOps", "P0", "DONE", "api, frontend, db, redis, minio, celery x2", "4h"],
        [10, "Set up PostgreSQL with RLS helpers", "Database", "P0", "DONE", "init.sql: uuid-ossp, pg_trgm, btree_gin", "2h"],
        [11, "Create Next.js 15 with RTL fonts + Tailwind v4", "Frontend", "P0", "DONE", "Inter + Noto Sans Arabic, globals.css", "3h"],
        [12, "Create login, register, forgot-password pages", "Frontend", "P0", "DONE", "Auth store (Zustand), API client (Axios)", "4h"],
    ]),
    2: ("CORE CRM", "W3-W4", [
        [1, "Create Contact, Tag, ContactTag models", "Backend", "P0", "DONE", "SoftDelete, search index, status index", "3h"],
        [2, "Create CustomField, CustomFieldValue models", "Backend", "P0", "DONE", "5 types: text/number/date/select/boolean", "2h"],
        [3, "Build contact service (CRUD, search, filter, sort, paginate)", "Backend", "P0", "DONE", "ilike search on name/phone/email, tag filter", "5h"],
        [4, "Build tag service (CRUD)", "Backend", "P0", "DONE", "Color-coded, company-unique names", "2h"],
        [5, "Build custom field service", "Backend", "P0", "DONE", "Sort ordering, field type validation", "2h"],
        [6, "Create contacts API (6 endpoints + bulk + import)", "Backend", "P0", "DONE", "/v1/contacts, /v1/tags, /v1/custom-fields", "4h"],
        [7, "Build CSV import Celery task", "Backend", "P1", "DONE", "Flexible column mapping, batch commits", "3h"],
        [8, "Build contacts list page with data table", "Frontend", "P0", "DONE", "Search, filters, sort, pagination, checkboxes", "5h"],
        [9, "Build contact detail/edit page", "Frontend", "P0", "DONE", "Inline edit, tag selection, custom fields", "4h"],
    ]),
    3: ("WHATSAPP INTEGRATION", "W5-W6", [
        [1, "Create Conversation model", "Backend", "P0", "DONE", "Status, assignment, last_msg cache, unread", "2h"],
        [2, "Create Message + MessageTemplate models", "Backend", "P0", "DONE", "Direction, delivery status, external_id", "3h"],
        [3, "Build WhatsApp Cloud API provider", "Backend", "P0", "DONE", "Send text/media/template/interactive/location", "5h"],
        [4, "Build webhook receiver (GET verify + POST events)", "Backend", "P0", "DONE", "Parse messages + status updates from Meta", "4h"],
        [5, "Build conversation service", "Backend", "P0", "DONE", "get_or_create, add_message, delivery tracking", "4h"],
        [6, "Build WebSocket manager (per-company broadcast)", "Backend", "P0", "DONE", "connect, disconnect, broadcast_to_company", "3h"],
        [7, "Create messaging Celery task", "Backend", "P0", "DONE", "Async send via Cloud API with retry", "2h"],
        [8, "Create conversation/message API (7 endpoints)", "Backend", "P0", "DONE", "List, detail, send, messages, update status", "4h"],
        [9, "Build shared inbox page (split-pane)", "Frontend", "P0", "DONE", "Conv list (left), chat thread (right)", "5h"],
        [10, "Add WebSocket endpoint to main.py", "Backend", "P0", "DONE", "/ws?token=jwt with ping/pong", "1h"],
    ]),
    4: ("SMART ROUTING & AUTOMATION", "W7-W8", [
        [1, "Build smart routing engine", "Backend", "P0", "DONE", "Relationship (40pts), skills (30), load (40), online (10)", "5h"],
        [2, "Create Automation, AutomationAction, AutomationLog models", "Backend", "P0", "DONE", "Trigger + conditions JSONB + actions ordered", "3h"],
        [3, "Build automation service (CRUD)", "Backend", "P0", "DONE", "Create, update, toggle, delete, logs", "3h"],
        [4, "Build condition evaluation engine", "Backend", "P0", "DONE", "7 operators: equals/contains/gt/lt/in/starts/not", "4h"],
        [5, "Build action execution engine", "Backend", "P0", "DONE", "8 types: auto_reply/tag/status/score/assign/template/webhook", "4h"],
        [6, "Create automation Celery task", "Backend", "P0", "DONE", "asyncio.run with fresh engine per task", "2h"],
        [7, "Create automation API (6 endpoints)", "Backend", "P0", "DONE", "CRUD + toggle + logs", "3h"],
        [8, "Integrate routing into webhook handler", "Backend", "P0", "DONE", "Auto-assign on new inbound conversation", "1h"],
        [9, "Build automation builder page", "Frontend", "P0", "DONE", "Trigger select, conditions, action config, logs panel", "4h"],
    ]),
    5: ("AI DIALECT ENGINE", "W9-W10", [
        [1, "Create AIConversationContext model", "Backend", "P0", "DONE", "Dialect, intent, sentiment, history JSONB, insights", "2h"],
        [2, "Build dialect detection (rule-based fallback)", "Backend", "P0", "DONE", "Arabic ratio + Kuwaiti markers (10+ words)", "3h"],
        [3, "Build intent classification (10 categories)", "Backend", "P0", "DONE", "pricing/purchase/support/complaint/greeting/...", "3h"],
        [4, "Build sentiment analysis", "Backend", "P0", "DONE", "positive/negative/neutral/mixed, -1.0 to 1.0", "2h"],
        [5, "Build Claude API integration", "Backend", "P1", "DONE", "System prompt with Kuwaiti dialect knowledge", "3h"],
        [6, "Build AI response generator", "Backend", "P1", "DONE", "Culturally appropriate, customer dialect match", "2h"],
        [7, "Create AI API (3 endpoints)", "Backend", "P0", "DONE", "POST /analyze, GET /context, POST /suggest", "2h"],
        [8, "Integrate AI into webhook flow (Celery)", "Backend", "P0", "DONE", "Auto-analyze every inbound message", "1h"],
    ]),
    6: ("SALES PIPELINE", "W11-W12", [
        [1, "Create Pipeline, PipelineStage models", "Backend", "P0", "DONE", "Ordered stages with color, is_won/is_lost", "2h"],
        [2, "Create Deal, DealActivity models", "Backend", "P0", "DONE", "Value DECIMAL(12,3), position, custom_data JSONB", "3h"],
        [3, "Build pipeline service (CRUD + stages)", "Backend", "P0", "DONE", "Add/remove stages, default pipeline", "3h"],
        [4, "Build deal service (CRUD + move + activities)", "Backend", "P0", "DONE", "Auto-log stage/value/status changes", "4h"],
        [5, "Build kanban board endpoint", "Backend", "P0", "DONE", "Deals grouped by stage with totals", "2h"],
        [6, "Create pipeline API (13 endpoints)", "Backend", "P0", "DONE", "Pipelines, stages, deals, board, activities, notes", "4h"],
        [7, "Build kanban board UI with drag-drop", "Frontend", "P0", "DONE", "Color columns, deal cards, HTML5 DnD", "5h"],
        [8, "Build create deal dialog", "Frontend", "P0", "DONE", "Title, value (KWD), stage selector", "2h"],
    ]),
    7: ("PAYMENTS & BILLING", "W13-W14", [
        [1, "Create Plan model (global)", "Backend", "P0", "DONE", "Pricing KWD, feature limits, flags", "2h"],
        [2, "Create Subscription, Invoice, Payment models", "Backend", "P0", "DONE", "Lifecycle, INV-YYYYMM-NNNN, Tap refs", "3h"],
        [3, "Build Tap Payments service", "Backend", "P0", "DONE", "K-Net/Visa/MC source mapping, 3D Secure", "4h"],
        [4, "Build subscription service", "Backend", "P0", "DONE", "Create, change plan, cancel, auto-invoice", "4h"],
        [5, "Build charge creation with redirect URL", "Backend", "P0", "DONE", "POST /charges returns payment_url", "2h"],
        [6, "Build Tap webhook handler", "Backend", "P0", "DONE", "CAPTURED/FAILED -> update payment + invoice", "2h"],
        [7, "Create payment API (10 endpoints)", "Backend", "P0", "DONE", "Plans, subscription, invoices, charges, webhook", "3h"],
        [8, "Seed default plans (Starter/Growth/Enterprise)", "Database", "P1", "DONE", "9.9/29.9/79.9 KWD monthly", "30m"],
        [9, "Build billing page with plan cards", "Frontend", "P0", "DONE", "Monthly/yearly toggle, invoice table", "4h"],
    ]),
    8: ("SHIPPING INTEGRATION", "W15-W16", [
        [1, "Create ShippingProvider, Shipment, TrackingEvent models", "Backend", "P0", "DONE", "Carrier config, COD, 7 status states", "3h"],
        [2, "Build abstract ShippingCarrier interface", "Backend", "P0", "DONE", "create_shipment, track, cancel", "2h"],
        [3, "Build Aramex Kuwait implementation", "Backend", "P0", "DONE", "Real API + mock mode for dev", "3h"],
        [4, "Build shipment service (CRUD + tracking)", "Backend", "P0", "DONE", "Auto-log events on status change", "4h"],
        [5, "Build carrier tracking refresh (poll API)", "Backend", "P1", "DONE", "POST /refresh polls carrier", "2h"],
        [6, "Build WhatsApp tracking notification task", "Backend", "P0", "DONE", "7 status message templates via Celery", "2h"],
        [7, "Create shipping API (8 endpoints)", "Backend", "P0", "DONE", "Providers, shipments, tracking, refresh", "3h"],
        [8, "Build shipment management page", "Frontend", "P1", "DONE", "List, status badges, tracking timeline", "3h"],
    ]),
    9: ("LANDING PAGES", "W17-W18", [
        [1, "Create LandingPage model", "Backend", "P0", "DONE", "Blocks JSONB, settings, WhatsApp CTA, SEO", "2h"],
        [2, "Build landing page service (CRUD + publish)", "Backend", "P0", "DONE", "Slug uniqueness, visit/conversion counters", "3h"],
        [3, "Build public page endpoint (no auth)", "Backend", "P0", "DONE", "GET /public/{slug} auto-increments visits", "2h"],
        [4, "Build conversion recording endpoint", "Backend", "P0", "DONE", "POST /public/{slug}/convert", "1h"],
        [5, "Build analytics endpoint", "Backend", "P1", "DONE", "Conversion rates per page", "1h"],
        [6, "Create landing pages API (10 endpoints)", "Backend", "P0", "DONE", "CRUD + publish + unpublish + public + convert", "3h"],
        [7, "Build page list with stats cards", "Frontend", "P0", "DONE", "Grid, status badges, visit/conversion rates", "3h"],
        [8, "Build block editor page", "Frontend", "P0", "DONE", "8 block types, add/remove/reorder, inline edit", "5h"],
    ]),
    10: ("ANALYTICS DASHBOARD", "W19-W20", [
        [1, "Build analytics service (dashboard stats)", "Backend", "P0", "DONE", "Contacts, conversations, messages, deals", "3h"],
        [2, "Build message volume stats (daily + delivery)", "Backend", "P0", "DONE", "GROUP BY date + direction", "2h"],
        [3, "Build pipeline stats (stages, win rate, avg deal)", "Backend", "P0", "DONE", "Outer join stages with deal aggregates", "2h"],
        [4, "Build team performance stats (per-agent)", "Backend", "P1", "DONE", "Messages sent, convs, deals won, revenue", "2h"],
        [5, "Build landing page + automation stats", "Backend", "P1", "DONE", "Visits, conversions, execution counts", "1h"],
        [6, "Build analytics dashboard page", "Frontend", "P0", "DONE", "KPI cards, pipeline chart, team table, LP stats", "5h"],
    ]),
    11: ("COMPLIANCE & AUDIT", "W21-W22", [
        [1, "Create AuditLog model (immutable)", "Backend", "P0", "DONE", "Who, what, resource, changes JSONB, IP", "2h"],
        [2, "Build audit service (log + query)", "Backend", "P0", "DONE", "Filter by action, resource, user", "2h"],
        [3, "Build compliance service (8-point checklist)", "Backend", "P0", "DONE", "Data residency, encryption, consent, retention", "3h"],
        [4, "Build compliance report generator", "Backend", "P0", "DONE", "Security measures, data summary, frameworks", "2h"],
        [5, "Create compliance API (3 endpoints)", "Backend", "P0", "DONE", "GET /status, /report, /audit-logs", "2h"],
        [6, "Build compliance dashboard page", "Frontend", "P0", "DONE", "Checklist, residency info, security list, logs", "4h"],
        [7, "Build audit log viewer", "Frontend", "P1", "DONE", "Table with action, description, timestamp", "2h"],
    ]),
    12: ("OPTIMIZATION & HARDENING", "W23-W24", [
        [1, "Build Redis caching layer", "Backend", "P0", "DONE", "get/set/delete/pattern, 5-min TTL, namespaced", "2h"],
        [2, "Build Prometheus metrics collector", "Backend", "P0", "DONE", "Counters, histograms (p50/p95/p99), gauges", "3h"],
        [3, "Integrate metrics into timing middleware", "Backend", "P0", "DONE", "http_requests_total, http_request_duration", "1h"],
        [4, "Create /metrics + /metrics/json endpoints", "Backend", "P0", "DONE", "Prometheus text + JSON format", "1h"],
        [5, "Implement 5 periodic Celery beat tasks", "Backend", "P0", "DONE", "analytics, payments, webhooks, auth, templates", "3h"],
        [6, "Fix Alembic + generate initial migration", "Database", "P0", "DONE", "alembic.ini, sys.path fix, stamp head", "2h"],
        [7, "Create .env file + fix docker-compose gaps", "DevOps", "P0", "DONE", "Nginx service, health checks, beat DB dep", "2h"],
    ]),
    13: ("PRODUCTION READINESS", "W25-W26", [
        [1, "Generate production secrets (JWT, app key, verify token)", "Security", "P0", "TODO", "openssl rand -hex 32", "1h"],
        [2, "Set WhatsApp Cloud API credentials", "Integration", "P0", "TODO", "Register at developers.facebook.com", "2h"],
        [3, "Set Tap Payments keys (K-Net, Visa, MC)", "Integration", "P0", "TODO", "tap.company developer portal", "2h"],
        [4, "Set Anthropic API key for AI engine", "Integration", "P1", "TODO", "console.anthropic.com", "30m"],
        [5, "Configure SSL/TLS (Let's Encrypt + nginx)", "Security", "P0", "TODO", "certbot, TLS 1.3, HSTS headers", "3h"],
        [6, "Add HTTPS + gzip + rate limit to nginx", "Security", "P0", "TODO", "ssl_certificate, gzip on, limit_req", "2h"],
        [7, "Create docker-compose.prod.yml", "DevOps", "P0", "TODO", "No volumes, env_file, built images, restart", "3h"],
        [8, "Set up AWS me-south-1 (EC2 or ECS Fargate)", "DevOps", "P0", "TODO", "Terraform or CloudFormation", "8h"],
        [9, "Configure RDS PostgreSQL 16 (multi-AZ)", "DevOps", "P0", "TODO", "Encrypted, automated backups, read replica", "4h"],
        [10, "Configure ElastiCache Redis", "DevOps", "P1", "TODO", "Cluster mode, encryption", "2h"],
        [11, "Set up S3 bucket (replace MinIO)", "DevOps", "P1", "TODO", "Versioning, lifecycle rules, CORS", "1h"],
        [12, "CI/CD deploy pipeline (GitHub Actions)", "DevOps", "P0", "TODO", "Build > push ECR > deploy ECS", "6h"],
        [13, "Add container registry push to CI", "DevOps", "P0", "TODO", "ECR or Docker Hub", "2h"],
        [14, "Configure Sentry error tracking", "Monitoring", "P1", "TODO", "Set SENTRY_DSN in production", "1h"],
        [15, "Set up CloudWatch log aggregation", "Monitoring", "P1", "TODO", "JSON structured logs ready", "4h"],
        [16, "Load test with k6 (100 concurrent users)", "Testing", "P1", "TODO", "Target <200ms p95 latency", "4h"],
        [17, "Database backup automation", "DevOps", "P0", "TODO", "RDS snapshots + pg_dump cron", "2h"],
        [18, "Domain DNS (app.kwgrowth.com)", "DevOps", "P0", "TODO", "A record + CNAME + SSL cert", "1h"],
    ]),
    14: ("UI POLISH & INTERNATIONALIZATION", "W27-W28", [
        [1, "Initialize shadcn/ui + install core components", "UI", "P0", "TODO", "npx shadcn-ui init; Button, Input, Select", "2h"],
        [2, "Add Table, Badge, Card, Dialog, Toast, DropdownMenu", "UI", "P0", "TODO", "Data display + interaction components", "2h"],
        [3, "Refactor contacts page with shadcn components", "UI", "P1", "TODO", "Replace inline Tailwind classes", "3h"],
        [4, "Refactor inbox with shadcn components", "UI", "P1", "TODO", "Conversation list, message bubbles, input", "3h"],
        [5, "Refactor pipeline kanban with @dnd-kit", "UI", "P1", "TODO", "DndContext, SortableContext, proper a11y", "4h"],
        [6, "Wire next-intl useTranslations() in all pages", "i18n", "P0", "TODO", "Replace 100+ hardcoded English strings", "4h"],
        [7, "Complete Arabic translations (common.json)", "i18n", "P0", "TODO", "All nav, forms, messages, statuses", "3h"],
        [8, "Add RTL layout toggle in settings", "i18n", "P1", "TODO", "dir=rtl on html, RTL Tailwind plugin", "2h"],
        [9, "Fix RTL layout issues (margins, padding, icons)", "i18n", "P1", "TODO", "Use logical properties ms/me vs ml/mr", "4h"],
        [10, "Responsive mobile layout for dashboard", "UI", "P1", "TODO", "Collapsible sidebar, hamburger menu", "4h"],
        [11, "Dark mode support", "UI", "P2", "TODO", "CSS variables ready, toggle + prefers-color-scheme", "3h"],
        [12, "Loading skeletons for all pages", "UX", "P2", "TODO", "Replace spinners with skeleton placeholders", "2h"],
        [13, "Toast notifications on all mutations", "UX", "P1", "TODO", "Success/error toasts on create/update/delete", "2h"],
        [14, "Form validation with Zod + react-hook-form", "UX", "P1", "TODO", "All forms: contacts, deals, automations", "3h"],
    ]),
    15: ("TESTING & QUALITY ASSURANCE", "W29-W30", [
        [1, "Unit tests: auth service (register, login, refresh)", "Backend", "P0", "TODO", "pytest + httpx AsyncClient", "3h"],
        [2, "Unit tests: contact service (CRUD, search, bulk)", "Backend", "P0", "TODO", "Mock DB with async fixtures", "3h"],
        [3, "Unit tests: conversation service + messages", "Backend", "P0", "TODO", "Message creation, status tracking", "3h"],
        [4, "Unit tests: automation engine (conditions + actions)", "Backend", "P0", "TODO", "Test all 7 operators, 8 action types", "4h"],
        [5, "Unit tests: AI dialect engine (fallback mode)", "Backend", "P1", "TODO", "Kuwaiti/Gulf/MSA/English detection", "2h"],
        [6, "Unit tests: pipeline service (deals, stages, board)", "Backend", "P1", "TODO", "Activity logging, kanban grouping", "3h"],
        [7, "Unit tests: Tap Payments (mock charges)", "Backend", "P1", "TODO", "Mock httpx, test webhook parsing", "2h"],
        [8, "Integration test: full auth flow E2E", "Backend", "P0", "TODO", "Register > login > me > refresh > logout", "2h"],
        [9, "Integration test: contact > conversation > message", "Backend", "P0", "TODO", "Create contact, start conv, send, check", "3h"],
        [10, "Integration test: webhook > automation pipeline", "Backend", "P1", "TODO", "Inbound msg triggers rule, executes actions", "3h"],
        [11, "Frontend component tests (Vitest + RTL)", "Frontend", "P1", "TODO", "Login form, contacts table, inbox", "4h"],
        [12, "Security audit: OWASP Top 10 checklist", "Security", "P0", "TODO", "SQLi, XSS, CSRF, broken auth, IDOR", "4h"],
    ]),
}

for snum in range(1, 16):
    title, weeks, tasks = all_sprint_tasks[snum]
    sheet_name = f"S{snum} - {title[:24]}"
    ws = wb.create_sheet(sheet_name)
    build_sprint_sheet(ws, snum, title, weeks, SPRINT_COLORS[snum], tasks)


# ══════════════════════════════════════════════════════════════════════════
# SHEET 17: Architecture
# ══════════════════════════════════════════════════════════════════════════
ws_arch = wb.create_sheet("Architecture")
ws_arch.sheet_properties.tabColor = "607D8B"
set_col_widths(ws_arch, [5, 28, 14, 24, 40])

title_row(ws_arch, "DATABASE SCHEMA - 29 TABLES", "607D8B", 5)
header_row(ws_arch, 3, ["#", "Table Name", "Module", "Mixins", "Description"])

tables = [
    ["users", "Auth", "Tenant+Timestamp", "JWT auth, profile, smart routing (skills, hours, online)"],
    ["roles", "Auth", "UUID+Timestamp", "5 system roles: platform_admin, owner, admin, manager, agent"],
    ["user_roles", "Auth", "Junction", "User-role many-to-many mapping"],
    ["api_keys", "Auth", "Tenant", "Scoped API keys with prefix, hash, expiration"],
    ["companies", "Core", "UUID+Timestamp", "Multi-tenant: name, slug, WhatsApp/Twilio/Tap encrypted config"],
    ["contacts", "CRM", "Tenant+SoftDelete", "Phone (E.164), email, lead_score, opt_in, source, assigned agent"],
    ["tags", "CRM", "Tenant", "Color-coded (#hex), company-unique name, description"],
    ["contact_tags", "CRM", "Junction", "Contact-tag many-to-many"],
    ["custom_fields", "CRM", "Tenant", "5 types: text, number, date, select, boolean; sort_order"],
    ["custom_field_values", "CRM", "Junction", "Contact-field values (text stored)"],
    ["conversations", "Messaging", "Tenant", "Status (open/closed/pending/snoozed), unread_count, last_msg cache"],
    ["messages", "Messaging", "Tenant", "Direction (in/out), sender_type, delivery_status, external_id"],
    ["message_templates", "Messaging", "Tenant", "WhatsApp templates: header, body, footer, buttons, status"],
    ["automations", "Automation", "Tenant", "Trigger event, conditions JSONB, priority, execution_count"],
    ["automation_actions", "Automation", "FK", "8 types: auto_reply/tag/status/score/assign/template/webhook"],
    ["automation_logs", "Automation", "Tenant", "Execution status, actions_executed, duration_ms, error"],
    ["ai_conversation_contexts", "AI", "Tenant", "Dialect, intent (+confidence), sentiment (+score), history JSONB"],
    ["pipelines", "Pipeline", "Tenant", "Named sales pipelines, is_default flag"],
    ["pipeline_stages", "Pipeline", "FK", "Colored stages, sort_order, is_won/is_lost terminal flags"],
    ["deals", "Pipeline", "Tenant+SoftDelete", "Value DECIMAL(12,3) KWD, status, position, expected_close"],
    ["deal_activities", "Pipeline", "FK", "Type: stage_changed/value_changed/status_changed/note_added"],
    ["plans", "Billing", "Global", "Pricing (monthly/yearly KWD), max_contacts/convs/team/automations"],
    ["subscriptions", "Billing", "Tenant", "Plan, cycle, period_start/end, cancel_at_period_end"],
    ["invoices", "Billing", "Tenant", "INV-YYYYMM-NNNN, subtotal/tax/total, line_items JSONB"],
    ["payments", "Billing", "Tenant", "Tap charge_id, payment_method (knet/visa/mc), card_last_four"],
    ["shipping_providers", "Shipping", "Tenant", "Carrier (aramex/dhl/fetchr/shipa), credentials, config"],
    ["shipments", "Shipping", "Tenant", "Tracking#, COD, addresses JSONB, 7 status states, notif tracking"],
    ["shipment_tracking_events", "Shipping", "FK", "Status, description, location, event_time, raw_data JSONB"],
    ["landing_pages", "Marketing", "Tenant+SoftDelete", "Blocks JSONB (8 types), slug, WhatsApp CTA, visit/conversion counters"],
    ["audit_logs", "Compliance", "Tenant", "Immutable: action, resource, changes JSONB, IP, user_agent"],
]

for idx, row in enumerate(tables):
    r = 4 + idx
    ws_arch.cell(row=r, column=1, value=idx+1).font = NORMAL
    ws_arch.cell(row=r, column=1).border = thin
    ws_arch.cell(row=r, column=1).alignment = ctr
    for c, val in enumerate(row, 2):
        cell = ws_arch.cell(row=r, column=c, value=val)
        cell.font = NORMAL
        cell.border = thin
        cell.alignment = lft if c >= 4 else ctr


# Save
output = r"D:\Work\Kuwait WhatsApp Growth Engine\App\KW_Growth_Engine_Sprint_Plan.xlsx"
wb.save(output)
print(f"Saved: {output}")
print(f"Sheets: {len(wb.sheetnames)}")
for s in wb.sheetnames:
    print(f"  - {s}")
